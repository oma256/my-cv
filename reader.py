from openpyxl import load_workbook


workbook = load_workbook('profile.xlsx')


def get_contact_info():
    worksheet = workbook['contact_info']
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data = {
            'fullname': row[0],
            'phone': row[1],
            'email': row[2],
            'linkedin': row[3],
            'location': row[4],
            'about_me': row[5]
        }

        return data


def get_education():
    worksheet = workbook['education']
    edu_list = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        edu_list.append({'name': row[0], 'description': row[1]})

    return edu_list


def get_data():
    data = {
        'contact_info': get_contact_info(),
        'education': get_education()
    }

    return data
